import openpyxl
import math
from openpyxl import Workbook
import pygame

class PyGameMaze:
    def __init__(self, colors, width, screen, laberinto):
        self.laberinto = laberinto
        self.cell_size = width // self.laberinto.shape[1]
        self.screen = screen
        self.colors = colors

    def draw_maze(self):
        filas, columnas = self.laberinto.shape
        
        for i in range(filas):
            for j in range(columnas):
                color = self.colors[self.laberinto[i][j]]
                pygame.draw.rect(self.screen, color, (j*self.cell_size, i*self.cell_size, self.cell_size, self.cell_size))
            
    def draw_pathcell(self,node,id_color):
        i, j = map(int, node.strip("()").split(","))

        color = self.colors[id_color]
        pygame.draw.rect(self.screen, color, (j*self.cell_size, i*self.cell_size, self.cell_size, self.cell_size))
        pygame.time.delay(10)
        pygame.display.flip()

def get_maze(file_path):
    matrix = []
    
    with open(file_path, 'r',encoding='utf-8') as file:
        for linea in file:
            matrix.append([int(char) for char in linea if char!="\n"])
            
    return matrix

def get_mazeStart(maze):
    filas, columnas = maze.shape
    for i in range(filas):
        for j in range(columnas):
            if maze[i][j]==2:
                return "("+str(i)+","+str(j)+")"
            
def get_mazeEnd(maze):
    filas, columnas = maze.shape
    for i in range(filas):
        for j in range(columnas):
            if maze[i][j]==3:
                return "("+str(i)+","+str(j)+")"

def create_costFuctionDoc(maze):
    # Crear un libro de trabajo
    wb = Workbook()
    # Seleccionar la hoja activa
    ws = wb.active
    
    # Puedes también añadir filas usando append
    filas, columnas = maze.shape

    ws.append(['Origen','Destino','Cost'])
    for i in range(filas):
        for j in range(columnas):
            posActual = maze[i][j]
            #Si es muro, omitir
            if posActual==0:
                continue
            
            #Izquierda
            if j>0:
                #Si es entrada, camino o meta, agregar
                if maze[i][j-1]!=0:
                    ws.append(["("+str(i)+","+str(j)+")","("+str(i)+","+str(j-1)+")",1])
            #Derecha
            if j<(columnas-1):
                #Si es entrada, camino o meta, agregar
                if maze[i][j+1]!=0:
                    ws.append(["("+str(i)+","+str(j)+")","("+str(i)+","+str(j+1)+")",1])
            #Arriba
            if i>0:
                #Si es entrada, camino o meta, agregar
                if maze[i-1][j]!=0:
                    ws.append(["("+str(i)+","+str(j)+")","("+str(i-1)+","+str(j)+")",1])
            #Abajo
            if i<(filas-1):
                #Si es entrada, camino o meta, agregar
                if maze[i+1][j]!=0:
                    ws.append(["("+str(i)+","+str(j)+")","("+str(i+1)+","+str(j)+")",1])

    # Guardar el libro de trabajo en un archivo
    wb.save("funcion_de_costo.xlsx")
    
def create_euclideanHeuristicDoc(maze):
    end = get_mazeEnd(maze)
    
    # Crear un libro de trabajo
    wb = Workbook()
    # Seleccionar la hoja activa
    ws = wb.active
    
    # Puedes también añadir filas usando append
    filas, columnas = maze.shape
    
    i2, j2 = map(int, end.strip("()").split(","))

    ws.append(['Origen','Distancia Euclediana'])
    for i1 in range(filas):
        for j1 in range(columnas):
            posActual = maze[i1][j1]
            #Si es muro, omitir
            if posActual==0:
                continue
            
            distancia = math.sqrt((i2 - i1)**2 + (j2 - j1)**2)
            ws.append(["("+str(i1)+","+str(j1)+")",distancia])
            
    # Guardar el libro de trabajo en un archivo
    wb.save("heuristica_euclediana.xlsx")
    
def create_manhattanHeuristicDoc(maze):
    end = get_mazeEnd(maze)
    
    # Crear un libro de trabajo
    wb = Workbook()
    # Seleccionar la hoja activa
    ws = wb.active
    
    # Puedes también añadir filas usando append
    filas, columnas = maze.shape
    
    i2, j2 = map(int, end.strip("()").split(","))

    ws.append(['Origen','Distancia Manhattan'])
    for i1 in range(filas):
        for j1 in range(columnas):
            posActual = maze[i1][j1]
            #Si es muro, omitir
            if posActual==0:
                continue
            
            distancia = abs(i2 - i1) + abs(j2 - j1)
            ws.append(["("+str(i1)+","+str(j1)+")",distancia])
            
    # Guardar el libro de trabajo en un archivo
    wb.save("heuristica_manhattan.xlsx")

def costFunction():
    archivo_excel = 'funcion_de_costo.xlsx'
    libro_trabajo = openpyxl.load_workbook(archivo_excel)
    hoja = libro_trabajo.active
    datos = []

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        datos.append(fila)
    
    libro_trabajo.close()
    
    return datos

def euclideanHeuristicFunction():
    archivo_excel = 'heuristica_euclediana.xlsx'
    libro_trabajo = openpyxl.load_workbook(archivo_excel)
    hoja = libro_trabajo.active
    datos = []

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        datos.append(fila)
    
    libro_trabajo.close()
    
    return datos

def manhattanHeuristicFunction():
    archivo_excel = 'heuristica_manhattan.xlsx'
    libro_trabajo = openpyxl.load_workbook(archivo_excel)
    hoja = libro_trabajo.active
    datos = []

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        datos.append(fila)
    
    libro_trabajo.close()
    
    return datos

def getGraph():
    graph = {}
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        graph[origen].append(destino)
        
    return graph

def getWeightedGraph():
    graph = {}
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = cost
        graph[origen].append(dic)
        
    return graph

def getEuclideanHeuristicTable():
    heuristicTable = {}
    for node, value in euclideanHeuristicFunction()[1:]:
        heuristicTable[node] = value
    
    return heuristicTable

def getManhattanHeuristicTable():
    heuristicTable = {}
    for node, value in manhattanHeuristicFunction()[1:]:
        heuristicTable[node] = value
    
    return heuristicTable

def getEuclideanHeuristicGraph():
    graph = {}
    heuristicTable = getEuclideanHeuristicTable()
        
    for origen, destino, _ in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = heuristicTable[destino]
        graph[origen].append(dic)
        
    return graph

def getManhattanHeuristicGraph():
    graph = {}
    heuristicTable = getManhattanHeuristicTable()
        
    for origen, destino, _ in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = heuristicTable[destino]
        graph[origen].append(dic)
        
    return graph

def getEuclideanCompleteGraph():
    graph = {}
    heuristicTable = getEuclideanHeuristicTable()
    
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = [cost,heuristicTable[destino]]
        graph[origen].append(dic)
        
    return graph

def getManhattanCompleteGraph():
    graph = {}
    heuristicTable = getManhattanHeuristicTable()
    
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = [cost,heuristicTable[destino]]
        graph[origen].append(dic)
        
    return graph